# GEMINI PROMPTED RESPONSE: This script performs a mass update on an Excel sheet based on a predefined set of updates. It uses the pandas library to read and manipulate the Excel file. The script identifies rows in the "current shipments" sheet where the "Client PO" matches the keys in the updates dictionary and updates the corresponding "horse reg" and "trailer reg" columns accordingly. Finally, it saves the updated data to a new Excel file.
# Make sure to install pandas if you haven't already:
# pip install pandas openpyxl
# 1. Define your update data
# This dictionary uses the "Client PO" as the key

import openpyxl

# 1. Define your update data
updates = {
    "BA3058": {"horse reg": "BCH9944ZM", "trailer reg": "BCE5712ZM / BCE5708ZM"},
    "BA3114": {"horse reg": "BCH9944ZM", "trailer reg": "BCE5712ZM / BCE5708ZM"},
    "BA3115": {"horse reg": "BCH9944ZM", "trailer reg": "BCE5712ZM / BCE5708ZM"},
    "BA3065": {"horse reg": "BCH9944ZM", "trailer reg": "BCE5712ZM / BCE5708ZM"},
    "BA3066": {"horse reg": "BCH9944ZM", "trailer reg": "BCE5712ZM / BCE5708ZM"},
    "BA3067": {"horse reg": "BCH9944ZM", "trailer reg": "BCE5712ZM / BCE5708ZM"},
    "BA3068": {"horse reg": "BCH9944ZM", "trailer reg": "BCE5712ZM / BCE5708ZM"},
}

file_path = r'C:\Users\Jason\Projects\TRACKING-FILE-UPDATE\BARTRAC - CONGO TRACKING UPD3 10-04-2026.xlsx'
target_sheet_name = 'current shipments'

wb = openpyxl.load_workbook(file_path)
found_sheet = next((s for s in wb.sheetnames if s.strip().lower() == target_sheet_name.lower()), None)

if not found_sheet:
    print(f"Error: Sheet '{target_sheet_name}' not found.")
    exit()

ws = wb[found_sheet]

# 2. Advanced Header Mapping (Handles 'horse reg no' and extra spaces)
header_map = {}
for cell in ws[1]:
    if cell.value:
        # Clean the header name: lowercase, strip spaces, remove internal double spaces
        clean_header = " ".join(str(cell.value).lower().split())
        header_map[clean_header] = cell.column

# Debugging: show what we mapped
print(f"Mapped 'client po' to column: {header_map.get('client po')}")
print(f"Mapped 'horse reg no' to column: {header_map.get('horse reg no')}")

# 3. Perform the update
update_count = 0

# Check for the correct column keys based on your output
col_po = header_map.get('client po')
col_horse = header_map.get('horse reg no')  # Updated to match your actual sheet
col_trailer = header_map.get('trailer reg no') # Updated to match your actual sheet

if not all([col_po, col_horse, col_trailer]):
    print("Error: Could not find required columns ('client po', 'horse reg no', or 'trailer reg no')")
    exit()

for row in range(2, ws.max_row + 1):
    raw_val = ws.cell(row=row, column=col_po).value
    if raw_val is None:
        continue
        
    po_val = str(raw_val).strip().upper()

    if po_val in updates:
        ws.cell(row=row, column=col_horse).value = updates[po_val]['horse reg']
        ws.cell(row=row, column=col_trailer).value = updates[po_val]['trailer reg']
        print(f"MATCH FOUND: Updated PO {po_val} at row {row}")
        update_count += 1

# 4. Save the file
output_path = r'C:\Users\Jason\Projects\TRACKING-FILE-UPDATE\BARTRAC_FINAL_UPDATED.xlsx'
wb.save(output_path)

print(f"\nUpdate Cycle Finished. Total Rows Updated: {update_count}")