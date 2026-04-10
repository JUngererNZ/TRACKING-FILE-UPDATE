"""
bartrac_update_vehicle_regs.py
-------------------------------
Mass-update HORSE REG NO and TRAILER REG NO on the CURRENT SHIPMENTS sheet
of a BARTRAC Congo Tracking workbook, matched by CLIENT PO.

Usage:
    python bartrac_update_vehicle_regs.py

Edit the UPDATES dict and INPUT_FILE / OUTPUT_FILE paths below, then run.
"""

from openpyxl import load_workbook
import sys

# ─────────────────────────────────────────────
#  CONFIG — edit these before running
# ─────────────────────────────────────────────

INPUT_FILE  = "BARTRAC_-_CONGO_TRACKING_10-04-2026.xlsx"
OUTPUT_FILE = "BARTRAC_-_CONGO_TRACKING_10-04-2026_UPDATED.xlsx"  # set same as INPUT_FILE to overwrite

SHEET_NAME  = "CURRENT SHIPMENTS "   # trailing space is intentional

# Add or remove entries as needed.
# Format: "CLIENT PO": {"horse": "HORSE REG VALUE", "trailer": "TRAILER REG VALUE"}
# Leave a key out (or set to None) to skip updating that field for a PO.
UPDATES = {
    "BA3075": {"horse": "ABC123ZM",  "trailer": "XYZ456ZM"},
    "BA3105": {"horse": "DEF789ZM",  "trailer": "GHI012ZM"},
    "BA3112": {"horse": "JKL345ZM",  "trailer": "MNO678ZM"},
    "BA3069": {"horse": "PQR901ZM",  "trailer": "STU234ZM"},
    "BA3070": {"horse": "VWX567ZM",  "trailer": "YZA890ZM"},
}

# ─────────────────────────────────────────────
#  CONSTANTS (match actual sheet layout)
# ─────────────────────────────────────────────

COL_CLIENT_PO  = 2   # Column B
COL_HORSE_REG  = 11  # Column K
COL_TRAILER_REG = 12 # Column L
HEADER_ROW     = 1


def run_update():
    print(f"Loading: {INPUT_FILE}")
    wb = load_workbook(INPUT_FILE)

    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    # Build lookup: PO value → row number (skip header)
    po_to_row = {}
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, min_col=COL_CLIENT_PO, max_col=COL_CLIENT_PO):
        cell = row[0]
        if cell.value:
            po_to_row[str(cell.value).strip()] = cell.row

    updated = []
    not_found = []

    for po, fields in UPDATES.items():
        po_clean = po.strip()
        if po_clean not in po_to_row:
            not_found.append(po_clean)
            continue

        row_num = po_to_row[po_clean]

        if fields.get("horse") is not None:
            ws.cell(row=row_num, column=COL_HORSE_REG).value = fields["horse"]

        if fields.get("trailer") is not None:
            ws.cell(row=row_num, column=COL_TRAILER_REG).value = fields["trailer"]

        updated.append(f"  {po_clean} (row {row_num}) → horse={fields.get('horse')}  trailer={fields.get('trailer')}")

    wb.save(OUTPUT_FILE)

    print(f"\n✓ Updated {len(updated)} PO(s):")
    for line in updated:
        print(line)

    if not_found:
        print(f"\n⚠ Not found in sheet ({len(not_found)}): {', '.join(not_found)}")

    print(f"\nSaved → {OUTPUT_FILE}")


if __name__ == "__main__":
    run_update()