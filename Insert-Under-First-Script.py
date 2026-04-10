import openpyxl
from copy import copy

file_path = r'C:\Users\Jason\Projects\TRACKING-FILE-UPDATE\BARTRAC - CONGO TRACKING UPD11 10-04-2026.xlsx'
target_sheet_name = 'current shipments'
target_horse = 'BCJ1334ZM'

wb = openpyxl.load_workbook(file_path)

# 1. Robust Sheet Finder
found_sheet_name = next((s for s in wb.sheetnames if s.strip().lower() == target_sheet_name.lower()), None)

if not found_sheet_name:
    print(f"Error: Sheet matching '{target_sheet_name}' not found.")
    exit()

ws = wb[found_sheet_name]
print(f"Successfully accessed sheet: '{found_sheet_name}'")

# 2. Map columns dynamically
header_map = {}
for cell in ws[1]:
    if cell.value:
        clean_header = " ".join(str(cell.value).lower().split())
        header_map[clean_header] = cell.column

col_horse = header_map.get('horse reg no')

# 3. Identify matches
all_match_indices = []
for row_idx in range(2, ws.max_row + 1):
    cell_val = str(ws.cell(row=row_idx, column=col_horse).value).strip()
    if cell_val == target_horse:
        all_match_indices.append(row_idx)

if len(all_match_indices) <= 1:
    print(f"Found {len(all_match_indices)} occurrence(s). No move required.")
    exit()

anchor_row_idx = all_match_indices[0]
rows_to_move_indices = sorted(all_match_indices[1:], reverse=True)

# 4. Extract data and styles using copy() to bypass StyleProxy issues
extracted_rows = []
for idx in rows_to_move_indices:
    row_data = [cell.value for cell in ws[idx]]
    # Use copy() on each style element to ensure they are unique instances
    row_styles = []
    for cell in ws[idx]:
        row_styles.append({
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format
        })
    extracted_rows.append((row_data, row_styles))
    ws.delete_rows(idx)

# 5. Insert rows under the anchor
for i, (data, styles) in enumerate(reversed(extracted_rows)):
    insert_pos = anchor_row_idx + 1 + i
    ws.insert_rows(insert_pos)
    for c_idx, value in enumerate(data, start=1):
        cell = ws.cell(row=insert_pos, column=c_idx, value=value)
        s = styles[c_idx-1]
        # Re-applying styles from our copied dictionary
        cell.font = s['font']
        cell.fill = s['fill']
        cell.border = s['border']
        cell.alignment = s['alignment']
        cell.number_format = s['number_format']

# 6. Save
output_path = r'C:\Users\Jason\Projects\TRACKING-FILE-UPDATE\BARTRAC_GROUPED_STABLE.xlsx'
wb.save(output_path)

print(f"\nSuccess! Grouped {len(all_match_indices)} rows for {target_horse}.")